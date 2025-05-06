import io
import contextlib
import sys
import os

from PIL import Image
from PyQt5.QtCore import QThread, pyqtSignal
from PyQt5.QtWidgets import (QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
                             QPushButton, QStackedWidget, QLabel, QFileDialog, QMessageBox,
                             QListWidget, QTextEdit, QLineEdit, QListWidgetItem, QGraphicsOpacityEffect, QScrollArea)
from PyQt5.QtCore import Qt, QSize, QEasingCurve, QRect, QUrl
from PyQt5.QtGui import QIcon, QFont, QColor
from PyQt5.QtMultimedia import QMediaPlayer, QMediaContent
from PyQt5.QtMultimediaWidgets import QVideoWidget
from pptx import Presentation
import pdfplumber
import pandas as pd
import json


class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("LittleToolkit")
        self.setGeometry(100, 100, 1200, 800)

        # 主窗口布局
        self.main_widget = QWidget()
        self.setCentralWidget(self.main_widget)

        self.main_layout = QHBoxLayout(self.main_widget)
        self.main_layout.setContentsMargins(0, 0, 0, 0)
        self.main_layout.setSpacing(0)

        # 创建侧边栏
        self.create_sidebar()

        # 创建主内容区域
        self.create_content_area()

        # 初始化各个功能页面
        self.create_home_page()
        self.create_tools_page()
        self.create_code_editor_page()
        self.create_app_store_page()

        # 默认显示主页
        self.stacked_widget.setCurrentIndex(0)
        # 初始化媒体播放器
        self.media_player = QMediaPlayer()
        self.video_widget = QVideoWidget()
        self.video_widget.hide()  # 默认隐藏视频窗口

        # 将视频窗口添加到主布局
        self.content_layout.addWidget(self.video_widget)
        self.media_player.setVideoOutput(self.video_widget)



    def create_sidebar(self):
        """创建侧边栏"""
        self.sidebar = QWidget()
        self.sidebar.setFixedWidth(250)
        self.sidebar.setStyleSheet("""
            background-color: #34495e;
            color: white;
        """)

        self.sidebar_layout = QVBoxLayout(self.sidebar)
        self.sidebar_layout.setContentsMargins(0, 20, 0, 20)
        self.sidebar_layout.setSpacing(20)

        # 添加侧边栏按钮
        self.btn_home = self.create_sidebar_button("主页", "home.png")
        self.btn_tools = self.create_sidebar_button("实用工具", "tools.png")
        self.btn_code_editor = self.create_sidebar_button("代码编辑器", "code.png")
        self.btn_app_store = self.create_sidebar_button("插件商店", "store.png")

        # 连接按钮信号
        self.btn_home.clicked.connect(lambda: self.stacked_widget.setCurrentIndex(0))
        self.btn_tools.clicked.connect(lambda: self.stacked_widget.setCurrentIndex(1))
        self.btn_code_editor.clicked.connect(lambda: self.stacked_widget.setCurrentIndex(2))
        self.btn_app_store.clicked.connect(lambda: self.stacked_widget.setCurrentIndex(3))

        # 添加按钮到侧边栏
        self.sidebar_layout.addWidget(self.btn_home)
        self.sidebar_layout.addWidget(self.btn_tools)
        self.sidebar_layout.addWidget(self.btn_code_editor)
        self.sidebar_layout.addWidget(self.btn_app_store)
        self.sidebar_layout.addStretch()

        # 将侧边栏添加到主布局
        self.main_layout.addWidget(self.sidebar)
        
    def load_plugin(self, plugin_path):
        """加载插件并创建侧边栏按钮
        Load plugin and create sidebar button
        
        参数:
            plugin_path: .ltc插件文件路径
        """
        try:
            # 检查文件扩展名
            if not plugin_path.endswith('.ltc'):
                print(f"不支持的插件格式: {plugin_path}")
                return
                
            # 使用utf-8编码读取插件文件
            with open(plugin_path, 'r', encoding='utf-8') as f:
                content = f.read()
                
            # 解析metadata部分
            metadata = {}
            code = ""
            in_metadata = False
            in_code = False
            
            for line in content.splitlines():
                if line.strip() == '[metadata]':
                    in_metadata = True
                    in_code = False
                elif line.strip() == '[code]':
                    in_metadata = False
                    in_code = True
                elif in_metadata and '=' in line:
                    key, value = line.split('=', 1)
                    metadata[key.strip()] = value.strip()
                elif in_code:
                    code += line + '\n'
            
            # 验证必填字段
            required_fields = ['name', 'version', 'description', 'category']
            for field in required_fields:
                if field not in metadata:
                    print(f"插件缺少必填字段: {field}")
                    return
            
            # 创建插件按钮
            if 'name' in metadata:
                # 检查是否已存在同名按钮
                for i in range(self.sidebar_layout.count()):
                    widget = self.sidebar_layout.itemAt(i).widget()
                    if isinstance(widget, QPushButton) and widget.text() == metadata['name']:
                        return
                        
                # 动态执行插件代码
                plugin_globals = {}
                exec(code, plugin_globals)
                
                # 获取main函数
                if 'main' in plugin_globals:
                    main_func = plugin_globals['main']
                    
                    # 创建按钮并绑定main函数
                    btn = self.create_sidebar_button(
                        metadata['name'], 
                        None, 
                        lambda: main_func()
                    )
                    self.sidebar_layout.insertWidget(self.sidebar_layout.count()-1, btn)
                else:
                    print("插件缺少main函数入口")
                    
        except Exception as e:
            print(f"加载插件失败: {str(e)}")

    def create_sidebar_button(self, text, icon_path=None, callback=None):
        """创建侧边栏按钮
        Create sidebar button
        """
        btn = QPushButton(text)
        btn.setFixedHeight(60)
        btn.setStyleSheet("""
            QPushButton {
                background-color: transparent;
                color: white;
                border: none;
                text-align: left;
                padding-left: 20px;
                font-size: 18px;
                border-radius: 10px;
            }
            QPushButton:hover {
                background-color: #2ecc71;
            }
        """)

        if icon_path:
            btn.setIcon(QIcon(icon_path))
            btn.setIconSize(QSize(32, 32))
            
        if callback:
            btn.clicked.connect(callback)

        return btn

    def create_content_area(self):
        """创建主内容区域"""
        self.content_widget = QWidget()
        self.content_widget.setStyleSheet("background-color: #ecf0f1;")

        self.content_layout = QVBoxLayout(self.content_widget)
        self.content_layout.setContentsMargins(30, 30, 30, 30)

        # 创建堆叠窗口用于切换不同页面
        self.stacked_widget = QStackedWidget()
        self.content_layout.addWidget(self.stacked_widget)

        # 添加终端输出区域
        self.terminal_output = QTextEdit()
        self.terminal_output.setStyleSheet("background-color: #f5f5f5; border-radius: 10px; height: 200px;")
        self.terminal_output.setReadOnly(True)

        # 将内容区域添加到主布局
        self.main_layout.addWidget(self.content_widget, 1)

    def create_home_page(self):
        """创建主页"""
        self.home_page = QWidget()
        layout = QVBoxLayout(self.home_page)

        # 标题
        title = QLabel("欢迎使用喵喵工具箱")
        title.setFont(QFont("Arial", 32, QFont.Bold))
        title.setAlignment(Qt.AlignCenter)
        layout.addWidget(title)

        # 搜索框
        search_layout = QHBoxLayout()
        self.search_input = QLineEdit()
        self.search_input.setPlaceholderText("输入搜索内容...")
        self.search_input.setStyleSheet("padding: 15px; font-size: 18px; border-radius: 10px;")
        self.search_input.setFixedHeight(50)

        search_btn = QPushButton("搜索")
        search_btn.setFixedSize(100, 50)
        search_btn.setStyleSheet("font-size: 18px; border-radius: 10px;")
        search_btn.clicked.connect(self.perform_search)

        search_layout.addWidget(self.search_input)
        search_layout.addWidget(search_btn)
        layout.addLayout(search_layout)

        # 搜索结果区域
        self.search_results = QTextEdit()
        self.search_results.setReadOnly(True)
        self.search_results.setStyleSheet("border-radius: 10px;")
        layout.addWidget(self.search_results)

        # 将主页添加到堆叠窗口
        self.stacked_widget.addWidget(self.home_page)

    def create_tools_page(self):
        """创建实用工具页面"""
        self.tools_page = QWidget()
        
        # 创建滚动区域
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll_content = QWidget()
        layout = QVBoxLayout(scroll_content)
        scroll.setWidget(scroll_content)
        
        self.tools_page.setLayout(QVBoxLayout())
        self.tools_page.layout().addWidget(scroll)
        
        # 添加终端输出区域到实用工具页面
        layout.addWidget(self.terminal_output)
        
        # PPT转PDF工具
        ppt_group = QWidget()
        ppt_layout = QVBoxLayout(ppt_group)

        ppt_title = QLabel("PPT转PDF工具")
        ppt_title.setFont(QFont("Arial", 20, QFont.Bold))
        ppt_layout.addWidget(ppt_title)

        ppt_instructions = QLabel("1. 选择PPT文件\n2. 点击转换按钮\n3. 等待转换完成")
        ppt_layout.addWidget(ppt_instructions)

        self.ppt_path_label = QLabel("未选择文件")
        ppt_layout.addWidget(self.ppt_path_label)

        btn_layout = QHBoxLayout()
        select_ppt_btn = QPushButton("选择PPT")
        select_ppt_btn.clicked.connect(self.select_ppt_file)
        convert_ppt_btn = QPushButton("转换")
        convert_ppt_btn.clicked.connect(self.convert_ppt_to_pdf)

        btn_layout.addWidget(select_ppt_btn)
        btn_layout.addWidget(convert_ppt_btn)
        ppt_layout.addLayout(btn_layout)

        layout.addWidget(ppt_group)

        # PDF转Excel工具
        pdf_excel_group = QWidget()
        pdf_excel_layout = QVBoxLayout(pdf_excel_group)

        pdf_excel_title = QLabel("PDF转Excel工具")
        pdf_excel_title.setFont(QFont("Arial", 20, QFont.Bold))
        pdf_excel_layout.addWidget(pdf_excel_title)

        pdf_excel_instructions = QLabel("1. 选择PDF文件\n2. 点击转换按钮\n3. 等待转换完成")
        pdf_excel_layout.addWidget(pdf_excel_instructions)

        self.pdf_excel_path_label = QLabel("未选择文件")
        pdf_excel_layout.addWidget(self.pdf_excel_path_label)

        btn_layout = QHBoxLayout()
        select_pdf_excel_btn = QPushButton("选择PDF")
        select_pdf_excel_btn.clicked.connect(self.select_pdf_excel_file)
        convert_pdf_excel_btn = QPushButton("转换")
        convert_pdf_excel_btn.clicked.connect(self.convert_pdf_to_excel)

        btn_layout.addWidget(select_pdf_excel_btn)
        btn_layout.addWidget(convert_pdf_excel_btn)
        pdf_excel_layout.addLayout(btn_layout)

        layout.addWidget(pdf_excel_group)
        
        # PDF转Word工具
        pdf_word_group = QWidget()
        pdf_word_layout = QVBoxLayout(pdf_word_group)

        pdf_word_title = QLabel("PDF转Word工具")
        pdf_word_title.setFont(QFont("Arial", 20, QFont.Bold))
        pdf_word_layout.addWidget(pdf_word_title)

        pdf_word_instructions = QLabel("1. 选择PDF文件\n2. 点击转换按钮\n3. 等待转换完成")
        pdf_word_layout.addWidget(pdf_word_instructions)

        self.pdf_word_path_label = QLabel("未选择文件")
        pdf_word_layout.addWidget(self.pdf_word_path_label)

        btn_layout = QHBoxLayout()
        select_pdf_word_btn = QPushButton("选择PDF")
        select_pdf_word_btn.clicked.connect(self.select_pdf_word_file)
        convert_pdf_word_btn = QPushButton("转换")
        convert_pdf_word_btn.clicked.connect(self.convert_pdf_to_word)

        btn_layout.addWidget(select_pdf_word_btn)
        btn_layout.addWidget(convert_pdf_word_btn)
        pdf_word_layout.addLayout(btn_layout)

        layout.addWidget(pdf_word_group)

        # PDF转Word工具
        pdf_group = QWidget()
        pdf_layout = QVBoxLayout(pdf_group)
        pdf_layout = QVBoxLayout(pdf_group)

        pdf_title = QLabel("PDF转Word工具")
        pdf_title.setFont(QFont("Arial", 20, QFont.Bold))
        pdf_layout.addWidget(pdf_title)

        pdf_instructions = QLabel("1. 选择PDF文件\n2. 点击转换按钮\n3. 等待转换完成")
        pdf_layout.addWidget(pdf_instructions)

        self.pdf_path_label = QLabel("未选择文件")
        pdf_layout.addWidget(self.pdf_path_label)

        btn_layout = QHBoxLayout()
        select_pdf_btn = QPushButton("选择PDF")
        select_pdf_btn.clicked.connect(self.select_pdf_file)
        convert_btn = QPushButton("转换")
        convert_btn.clicked.connect(self.convert_pdf_to_word)

        btn_layout.addWidget(select_pdf_btn)
        btn_layout.addWidget(convert_btn)
        pdf_layout.addLayout(btn_layout)

        layout.addWidget(pdf_group)

        # Excel分表工具
        excel_group = QWidget()
        excel_layout = QVBoxLayout(excel_group)

        excel_title = QLabel("Excel分表工具")
        excel_title.setFont(QFont("Arial", 20, QFont.Bold))
        excel_layout.addWidget(excel_title)

        excel_instructions = QLabel("1. 选择Excel文件\n2. 选择输出目录\n3. 点击拆分按钮")
        excel_layout.addWidget(excel_instructions)

        self.excel_path_label = QLabel("未选择文件")
        excel_layout.addWidget(self.excel_path_label)

        self.excel_output_label = QLabel("未选择输出目录")
        excel_layout.addWidget(self.excel_output_label)

        excel_btn_layout = QHBoxLayout()
        select_excel_btn = QPushButton("选择Excel")
        select_excel_btn.clicked.connect(self.select_excel_file)
        select_output_btn = QPushButton("选择输出目录")
        select_output_btn.clicked.connect(self.select_output_folder)
        split_btn = QPushButton("拆分")
        split_btn.clicked.connect(self.split_excel_sheets)

        excel_btn_layout.addWidget(select_excel_btn)
        excel_btn_layout.addWidget(select_output_btn)
        excel_btn_layout.addWidget(split_btn)
        excel_layout.addLayout(excel_btn_layout)

        layout.addWidget(excel_group)
        
        # GIF拆分工具
        gif_group = QWidget()
        gif_layout = QVBoxLayout(gif_group)

        gif_title = QLabel("GIF拆分工具")
        gif_title.setFont(QFont("Arial", 20, QFont.Bold))
        gif_layout.addWidget(gif_title)

        gif_instructions = QLabel("1. 选择GIF文件\n2. 选择输出目录\n3. 点击拆分按钮")
        gif_layout.addWidget(gif_instructions)

        self.gif_path_label = QLabel("未选择文件")
        gif_layout.addWidget(self.gif_path_label)

        self.gif_output_label = QLabel("未选择输出目录")
        gif_layout.addWidget(self.gif_output_label)

        gif_btn_layout = QHBoxLayout()
        select_gif_btn = QPushButton("选择GIF")
        select_gif_btn.clicked.connect(self.select_gif_file)
        select_gif_output_btn = QPushButton("选择输出目录")
        select_gif_output_btn.clicked.connect(self.select_gif_output_folder)
        split_gif_btn = QPushButton("拆分")
        split_gif_btn.clicked.connect(self.split_gif_frames)

        gif_btn_layout.addWidget(select_gif_btn)
        gif_btn_layout.addWidget(select_gif_output_btn)
        gif_btn_layout.addWidget(split_gif_btn)
        gif_layout.addLayout(gif_btn_layout)

        layout.addWidget(gif_group)

        # GIF合并工具
        gif_merge_group = QWidget()
        gif_merge_layout = QVBoxLayout(gif_merge_group)

        gif_merge_title = QLabel("GIF合并工具")
        gif_merge_title.setFont(QFont("Arial", 20, QFont.Bold))
        gif_merge_layout.addWidget(gif_merge_title)

        gif_merge_instructions = QLabel("1. 选择图片\n2. 设置帧间隔\n3. 点击合并按钮")
        gif_merge_layout.addWidget(gif_merge_instructions)

        self.gif_merge_status_label = QLabel("未选择图片")
        gif_merge_layout.addWidget(self.gif_merge_status_label)

        self.gif_merge_interval_input = QLineEdit("100")
        self.gif_merge_interval_input.setPlaceholderText("帧间隔 (ms)")
        gif_merge_layout.addWidget(self.gif_merge_interval_input)

        gif_merge_btn_layout = QHBoxLayout()
        select_images_btn = QPushButton("选择图片")
        select_images_btn.clicked.connect(self.select_images_for_gif)
        merge_gif_btn = QPushButton("合并")
        merge_gif_btn.clicked.connect(self.merge_images_to_gif)

        gif_merge_btn_layout.addWidget(select_images_btn)
        gif_merge_btn_layout.addWidget(merge_gif_btn)
        gif_merge_layout.addLayout(gif_merge_btn_layout)

        layout.addWidget(gif_merge_group)
        layout.addStretch()

        self.stacked_widget.addWidget(self.tools_page)

    def create_code_editor_page(self):
        """创建代码编辑器页面"""
        self.code_editor_page = QWidget()
        layout = QVBoxLayout(self.code_editor_page)

        # 代码编辑区域
        self.code_editor = QTextEdit()
        self.code_editor.setFont(QFont("Consolas", 14))
        self.code_editor.setStyleSheet("background-color: white; border-radius: 10px;")
        layout.addWidget(self.code_editor)

        # 按钮区域
        btn_layout = QHBoxLayout()
        run_btn = QPushButton("运行")
        run_btn.clicked.connect(self.run_code)
        save_btn = QPushButton("保存")
        save_btn.clicked.connect(self.save_code)
        clear_btn = QPushButton("清空")
        clear_btn.clicked.connect(self.clear_code)

        btn_layout.addWidget(run_btn)
        btn_layout.addWidget(save_btn)
        btn_layout.addWidget(clear_btn)
        layout.addLayout(btn_layout)

        # 输出区域
        self.code_output = QTextEdit()
        self.code_output.setReadOnly(True)
        self.code_output.setStyleSheet("background-color: #f5f5f5; border-radius: 10px;")
        layout.addWidget(self.code_output)

        self.stacked_widget.addWidget(self.code_editor_page)

    def create_app_store_page(self):
        """创建插件商店页面
        Create app store page
        """
        self.app_store_page = QWidget()
        
        layout = QVBoxLayout(self.app_store_page)
        
        # 计算器插件安装按钮
        calculator_btn = QPushButton("安装计算器插件")
        calculator_btn.setStyleSheet("""
            QPushButton {
                background-color: #2ecc71;
                color: white;
                border: none;
                padding: 15px;
                font-size: 18px;
                border-radius: 10px;
            }
            QPushButton:hover {
                background-color: #27ae60;
            }
        """)
        calculator_btn.clicked.connect(lambda: self.load_plugin("numberbot.ltc"))
        
        layout.addWidget(calculator_btn)
        layout.addStretch()
        layout = QHBoxLayout(self.app_store_page)

        # 插件列表
        self.app_list = QListWidget()
        self.app_list.setFixedWidth(350)
        self.app_list.itemClicked.connect(self.show_app_details)
        self.app_list.setStyleSheet("border-radius: 10px;")

        # 模拟插件数据
        apps = [
            {"name": "文本编辑器", "description": "一个简单的文本编辑工具", "category": "工具", "file": "text_editor.ltc"},
            {"name": "计算器", "description": "科学计算器", "category": "工具", "file": "calculator.ltc"},
            {"name": "天气插件", "description": "查看实时天气信息", "category": "实用程序", "file": "weather.ltc"},
            {"name": "神秘彩蛋", "description": "点击这里看看惊喜！", "category": "秘密", "file": "easter_egg.ltc"}
        ]

        for app in apps:
            item = QListWidgetItem(app["name"])
            item.setData(Qt.UserRole, app)
            self.app_list.addItem(item)

        # 应用详情
        self.app_details = QWidget()
        details_layout = QVBoxLayout(self.app_details)

        self.app_title = QLabel()
        self.app_title.setFont(QFont("Arial", 22, QFont.Bold))
        details_layout.addWidget(self.app_title)

        self.app_description = QLabel()
        self.app_description.setWordWrap(True)
        details_layout.addWidget(self.app_description)

        self.app_category = QLabel()
        details_layout.addWidget(self.app_category)

        install_btn = QPushButton("安装插件")
        install_btn.clicked.connect(self.install_plugin)
        details_layout.addWidget(install_btn)
        details_layout.addStretch()

        layout.addWidget(self.app_list)
        layout.addWidget(self.app_details)

        self.stacked_widget.addWidget(self.app_store_page)

    # 以下是各个功能的实现方法
    def perform_search(self):
        """执行搜索功能"""
        query = self.search_input.text()
        if not query:
            QMessageBox.warning(self, "警告", "请输入搜索内容")
            return

        self.search_results.setText(f"正在搜索: {query}\n\n")
        # 这里可以添加实际的搜索逻辑
        self.search_results.append("搜索结果1: 示例文件1.txt")
        self.search_results.append("搜索结果2: 示例文件2.pdf")

    def select_pdf_file(self):
        """选择PDF文件"""
        file_path, _ = QFileDialog.getOpenFileName(self, "选择PDF文件", "", "PDF文件 (*.pdf)")
        if file_path:
            self.pdf_path_label.setText(file_path)

    def select_ppt_file(self):
        """
        选择PPT文件
        Select PPT file
        """
        file_path, _ = QFileDialog.getOpenFileName(
            self, "选择PPT文件", "", "PPT Files (*.pptx *.ppt)")
        if file_path:
            self.ppt_path = file_path
            self.ppt_path_label.setText(file_path)

    def convert_ppt_to_pdf(self):
        """
        将PPT文件转换为PDF
        Convert PPT to PDF
        """
        if not hasattr(self, 'ppt_path'):
            QMessageBox.warning(self, "警告", "请先选择PPT文件！")
            return

        try:
            output_path = self.ppt_path.replace('.pptx', '.pdf').replace('.ppt', '.pdf')
            prs = Presentation(self.ppt_path)
            prs.save(output_path)
            QMessageBox.information(self, "成功", f"文件已转换为 {output_path}")
        except Exception as e:
            QMessageBox.critical(self, "错误", f"转换失败: {str(e)}")

    def select_pdf_excel_file(self):
        """
        选择PDF文件(用于转Excel)
        Select PDF file (for Excel conversion)
        """
        file_path, _ = QFileDialog.getOpenFileName(
            self, "选择PDF文件", "", "PDF Files (*.pdf)")
        if file_path:
            self.pdf_excel_path = file_path
            self.pdf_excel_path_label.setText(file_path)
            
    def select_pdf_word_file(self):
        """
        选择PDF文件(用于转Word)
        Select PDF file (for Word conversion)
        """
        file_path, _ = QFileDialog.getOpenFileName(
            self, "选择PDF文件", "", "PDF Files (*.pdf)")
        if file_path:
            self.pdf_path = file_path
            self.pdf_word_path_label.setText(file_path)

    def convert_pdf_to_excel(self):
        """
        将PDF表格数据转换为Excel
        Convert PDF tables to Excel
        """
        if not hasattr(self, 'pdf_excel_path'):
            QMessageBox.warning(self, "警告", "请先选择PDF文件！")
            return

        try:
            output_path = self.pdf_excel_path.replace('.pdf', '.xlsx')
            with pdfplumber.open(self.pdf_excel_path) as pdf_file:
                first_page = pdf_file.pages[0]
                tables = first_page.extract_tables()
                if tables:
                    df = pd.DataFrame(tables[0])
                    df.to_excel(output_path, index=False)
                    QMessageBox.information(self, "成功", f"文件已转换为 {output_path}")
                else:
                    QMessageBox.warning(self, "警告", "未找到表格数据！")
        except Exception as e:
            QMessageBox.critical(self, "错误", f"转换失败: {str(e)}")

    def convert_pdf_to_word(self):
        """
        将PDF文件转换为Word文件
        Convert PDF file to Word document
        """
        if not hasattr(self, 'pdf_path'):
            QMessageBox.warning(self, "警告", "请先选择PDF文件！")
            self.terminal_output.append("错误: 未选择PDF文件\nError: No PDF file selected")
            return

        if not os.path.exists(self.pdf_path):
            QMessageBox.warning(self, "警告", "PDF文件不存在！")
            self.terminal_output.append(f"错误: 文件不存在 {self.pdf_path}\nError: File not found {self.pdf_path}")
            return

        try:
            self.terminal_output.append(f"开始转换: {self.pdf_path}\nStarting conversion: {self.pdf_path}")
            output_path = self.pdf_path.replace('.pdf', '.docx')
            
            # 检查输出路径是否可写
            try:
                with open(output_path, 'w') as f:
                    pass
                os.remove(output_path)
            except Exception as e:
                QMessageBox.critical(self, "错误", f"输出路径不可写: {str(e)}")
                self.terminal_output.append(f"错误: 输出路径不可写 {output_path}\nError: Output path not writable {output_path}")
                return

            # 确保已安装pdf2docx库
            try:
                from pdf2docx import Converter
            except ImportError:
                QMessageBox.critical(self, "错误", "请先安装pdf2docx库: pip install pdf2docx -i https://mirrors.aliyun.com/pypi/simple/")
                self.terminal_output.append("错误: 未安装pdf2docx库\nError: pdf2docx library not installed")
                return

            # 调用Converter进行转换
            cv = Converter(self.pdf_path)
            cv.convert(output_path, start=0, end=None)
            cv.close()
            
            self.terminal_output.append(f"转换成功: {output_path}\nConversion successful: {output_path}")
            QMessageBox.information(self, "成功", f"文件已转换为 {output_path}")
            
        except Exception as e:
            error_msg = f"转换失败: {str(e)}\nConversion failed: {str(e)}"
            self.terminal_output.append(error_msg)
            QMessageBox.critical(self, "错误", error_msg)

    def select_excel_file(self):
        """选择Excel文件"""
        file_path, _ = QFileDialog.getOpenFileName(self, "选择Excel文件", "", "Excel文件 (*.xlsx *.xls)")
        if file_path:
            self.excel_path_label.setText(file_path)

    def select_output_folder(self):
        """选择输出目录"""
        folder_path = QFileDialog.getExistingDirectory(self, "选择输出目录")
        if folder_path:
            self.excel_output_label.setText(folder_path)

    def select_gif_file(self):
        """选择GIF文件
        Select GIF file
        """
        file_path, _ = QFileDialog.getOpenFileName(self, "选择GIF文件", "", "GIF文件 (*.gif)")
        if file_path:
            self.gif_path_label.setText(file_path)

    def select_gif_output_folder(self):
        """选择GIF输出目录
        Select GIF output folder
        """
        folder_path = QFileDialog.getExistingDirectory(self, "选择GIF输出目录")
        if folder_path:
            self.gif_output_label.setText(folder_path)

    def split_gif_frames(self):
        """拆分GIF帧
        Split GIF frames
        """
        gif_path = self.gif_path_label.text()
        output_path = self.gif_output_label.text()

        if gif_path == "未选择文件":
            QMessageBox.warning(self, "警告", "请先选择GIF文件")
            return
        if output_path == "未选择输出目录":
            QMessageBox.warning(self, "警告", "请先选择输出目录")
            return

        try:
            from PIL import Image
            
            # 打开GIF文件
            gif = Image.open(gif_path)
            
            # 遍历每一帧并保存
            for i in range(gif.n_frames):
                gif.seek(i)
                frame_path = os.path.join(output_path, f"frame_{i}.png")
                gif.save(frame_path)
                self.terminal_output.append(f"已保存第 {i} 帧到 {frame_path}")
            
            self.terminal_output.append(f"\n拆分完成，共保存了 {gif.n_frames} 帧")
        except Exception as e:
            self.terminal_output.append(f"\n拆分GIF失败: {str(e)}")

    def split_excel_sheets(self):
        """拆分Excel工作表
        Split Excel sheets
        """
        excel_path = self.excel_path_label.text()
        output_path = self.excel_output_label.text()

        if excel_path == "未选择文件":
            QMessageBox.warning(self, "警告", "请先选择Excel文件")
            return
        if output_path == "未选择输出目录":
            QMessageBox.warning(self, "警告", "请先选择输出目录")
            return

        try:
            # 根据文件扩展名选择不同的处理方式
            if excel_path.lower().endswith('.xlsx'):
                import openpyxl
                # 读取.xlsx文件
                wb = openpyxl.load_workbook(excel_path)
                
                # 遍历每个工作表并保存为单独文件
                for sheet_name in wb.sheetnames:
                    try:
                        # 创建新工作簿并复制当前工作表
                        new_wb = openpyxl.Workbook()
                        new_sheet = new_wb.active
                        
                        # 复制数据
                        for row in wb[sheet_name].iter_rows():
                            new_row = []
                            for cell in row:
                                new_row.append(cell.value)
                            new_sheet.append(new_row)
                        
                        # 保存文件
                        output_file = os.path.join(output_path, f"{sheet_name}.xlsx")
                        new_wb.save(output_file)
                        self.terminal_output.append(f"已保存工作表 {sheet_name} 到 {output_file}")
                    except Exception as e:
                        self.terminal_output.append(f"保存工作表 {sheet_name} 失败: {str(e)}")
            elif excel_path.lower().endswith('.xls'):
                import xlrd
                # 读取.xls文件
                wb = xlrd.open_workbook(excel_path)
                
                # 遍历每个工作表并保存为单独文件
                for sheet_name in wb.sheet_names():
                    try:
                        # 创建新工作簿
                        import openpyxl
                        new_wb = openpyxl.Workbook()
                        new_sheet = new_wb.active
                        
                        # 读取数据
                        sheet = wb.sheet_by_name(sheet_name)
                        for row_idx in range(sheet.nrows):
                            new_row = sheet.row_values(row_idx)
                            new_sheet.append(new_row)
                        
                        # 保存文件
                        output_file = os.path.join(output_path, f"{sheet_name}.xlsx")
                        new_wb.save(output_file)
                        self.terminal_output.append(f"已保存工作表 {sheet_name} 到 {output_file}")
                    except Exception as e:
                        self.terminal_output.append(f"保存工作表 {sheet_name} 失败: {str(e)}")
            else:
                raise ValueError("不支持的文件格式，请使用.xls或.xlsx文件")
            
            QMessageBox.information(self, "成功", f"已拆分Excel文件到: {output_path}")
        except Exception as e:
            QMessageBox.critical(self, "错误", f"拆分Excel文件失败: {str(e)}")

    def run_code(self):
        """运行代码"""
        code = self.code_editor.toPlainText()
        if not code:
            QMessageBox.warning(self, "警告", "请输入代码")
            return

        # 重定向标准输出
        old_stdout = sys.stdout
        redirected_output = sys.stdout = io.StringIO()

        try:
            # 执行代码并捕获输出
            with contextlib.redirect_stdout(redirected_output):
                exec(code)

            # 获取执行结果
            output = redirected_output.getvalue()
            self.code_output.setText("执行结果:\n" + output)

        except Exception as e:
            self.code_output.setText(f"执行错误:\n{str(e)}")
        finally:
            sys.stdout = old_stdout

    def save_code(self):
        """保存代码"""
        code = self.code_editor.toPlainText()
        if not code:
            QMessageBox.warning(self, "警告", "没有代码可保存")
            return

        file_path, _ = QFileDialog.getSaveFileName(self, "保存代码", "", "Python文件 (*.py);;所有文件 (*)")
        if file_path:
            try:
                with open(file_path, 'w') as f:
                    f.write(code)
                QMessageBox.information(self, "成功", f"代码已保存到: {file_path}")
            except Exception as e:
                QMessageBox.critical(self, "错误", f"保存失败: {str(e)}")

    def clear_code(self):
        """清空代码"""
        self.code_editor.clear()
        self.code_output.clear()

    def show_app_details(self, item):
        """显示应用详情"""
        app_data = item.data(Qt.UserRole)
        self.app_title.setText(app_data["name"])
        self.app_description.setText(app_data["description"])
        self.app_category.setText(f"类别: {app_data['category']}")

        if app_data["name"] == "神秘彩蛋":
            self.play_video()

    def play_video(self):
        """播放视频"""
        try:
            base_dir = os.path.dirname(os.path.abspath(__file__))
            video_file = "AI骗人.mp4"
            video_url = os.path.join(base_dir, video_file)

            print(f"视频路径: {video_url}")
            print(f"文件存在: {os.path.exists(video_url)}")

            if not os.path.exists(video_url):
                QMessageBox.warning(self, "警告", f"视频文件不存在: {video_url}")
                return

            # 检查媒体支持
            if not self.media_player.isAvailable():
                QMessageBox.critical(self, "错误", "媒体服务不可用")
                return

            print(f"支持的MIME类型: {self.media_player.supportedMimeTypes()}")

            # 尝试设置媒体内容
            media_content = QMediaContent(QUrl.fromLocalFile(video_url))
            self.media_player.setMedia(media_content)

            # 连接信号
            self.media_player.error.connect(self.handle_media_error)
            self.media_player.mediaStatusChanged.connect(self.handle_media_status)

            self.video_widget.show()
            self.media_player.play()

        except Exception as e:
            QMessageBox.critical(self, "错误", f"播放失败: {str(e)}")

    def handle_media_status(self, status):
        """处理媒体状态变化"""
        print(f"媒体状态: {status}")
        if status == QMediaPlayer.LoadedMedia:
            print("媒体已加载")
        elif status == QMediaPlayer.InvalidMedia:
            print("无效媒体格式")

    def handle_media_error(self, error):
        """处理媒体播放错误"""
        error_msg = self.media_player.errorString()
        print(f"媒体播放错误: {error_msg}")
        QMessageBox.critical(self, "播放错误", f"无法播放视频: {error_msg}")

    def install_plugin(self):
        """安装插件"""
        current_item = self.app_list.currentItem()
        if current_item:
            plugin_data = current_item.data(Qt.UserRole)
            try:
                # 解析.ltc插件文件
                with open(plugin_data['file'], 'r', encoding='utf-8') as f:
                    plugin_content = f.read()
                    # TODO: 添加实际插件安装逻辑
                    QMessageBox.information(self, "安装成功", 
                        f"已成功安装插件 {plugin_data['name']}\n文件: {plugin_data['file']}")
            except Exception as e:
                QMessageBox.critical(self, "安装失败", f"插件安装失败: {str(e)}")

    def perform_search(self):
        """执行搜索功能"""
        query = self.search_input.text().strip()
        if not query:
            QMessageBox.warning(self, "警告", "请输入搜索内容")
            return

        self.search_results.setText(f"正在搜索: {query}\n\n")

        # 创建并启动工作线程
        self.search_thread = SearchWorker(query, os.getcwd())
        self.search_thread.found_match.connect(self.search_results.append)
        self.search_thread.finished.connect(lambda: self.search_results.append("\n搜索完成！"))
        self.search_thread.start()

    def select_images_for_gif(self):
        """选择图片文件用于GIF合并"""
        file_paths, _ = QFileDialog.getOpenFileNames(self, "选择图片文件", "", "图片文件 (*.png *.jpg *.jpeg)")
        if file_paths:
            self.gif_merge_status_label.setText(f"已选择 {len(file_paths)} 张图片")
            self.selected_images = file_paths

    def merge_images_to_gif(self):
        """合并图片为GIF"""
        try:
            interval = int(self.gif_merge_interval_input.text())
            if interval <= 0:
                QMessageBox.warning(self, "警告", "帧间隔必须为正整数")
                return

            if not hasattr(self, 'selected_images') or not self.selected_images:
                QMessageBox.warning(self, "警告", "请先选择图片")
                return

            output_path, _ = QFileDialog.getSaveFileName(self, "保存GIF文件", "", "GIF文件 (*.gif)")
            if not output_path:
                return

            # 打开所有图片
            images = [Image.open(path) for path in self.selected_images]

            # 合并图片为GIF
            images[0].save(output_path, save_all=True, append_images=images[1:], loop=0, duration=interval)
            QMessageBox.information(self, "成功", f"GIF已保存到: {output_path}")
        except Exception as e:
            QMessageBox.critical(self, "错误", f"合并失败: {str(e)}")


class SearchWorker(QThread):
    found_match = pyqtSignal(str)  # 信号，用于发送找到的匹配项

    def __init__(self, query, search_dir):
        super().__init__()
        self.query = query.lower()
        self.search_dir = search_dir

    def run(self):
        """执行搜索的线程方法"""
        try:
            for root, dirs, files in os.walk(self.search_dir):
                for file in files:
                    file_path = os.path.join(root, file)

                    # 检查文件名匹配
                    if self.query in file.lower():
                        self.found_match.emit(f"文件名匹配: {file_path}")

                    # 检查文件内容匹配
                    if file.endswith(('.txt', '.py', '.md', '.html', '.js', '.css')):
                        try:
                            with open(file_path, 'r', encoding='utf-8') as f:
                                for line in f:
                                    if self.query in line.lower():
                                        self.found_match.emit(f"内容匹配: {file_path}")
                                        break
                        except (UnicodeDecodeError, PermissionError):
                            continue
        except Exception as e:
            self.found_match.emit(f"搜索出错: {str(e)}")

if __name__ == "__main__":
    app = QApplication(sys.argv)

    # 设置应用程序样式
    app.setStyle("Fusion")

    window = MainWindow()
    window.show()
    sys.exit(app.exec_())


